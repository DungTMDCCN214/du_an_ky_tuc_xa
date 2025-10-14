# dormitory/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Room, Building, Contract, Student

def home(request):
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    total_buildings = Building.objects.count()
    active_contracts = Contract.objects.filter(status='active').count()
    
    context = {
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'total_buildings': total_buildings,
        'active_contracts': active_contracts,
    }
    return render(request, 'dormitory/home.html', context)


@login_required
def student_dashboard(request):
    # Dashboard cho sinh vi√™n
    if request.user.user_type != 'student':
        messages.error(request, "B·∫°n kh√¥ng c√≥ quy·ªÅn truy c·∫≠p trang n√†y")
        return redirect('home')
    
    try:
        student = Student.objects.get(user=request.user)
        current_contract = Contract.objects.filter(
            student=student, 
            status='active'
        ).first()
        
        # L·∫§Y DANH S√ÅCH PH√íNG C√íN CH·ªñ TR·ªêNG
        available_rooms = Room.objects.filter(
            status='available'
        ).select_related('building', 'room_type')
        
        # L·ªåC CH·ªà NH·ªÆNG PH√íNG C√íN CH·ªñ TR·ªêNG
        available_rooms = [room for room in available_rooms if room.is_available()]
        
        context = {
            'student': student,
            'current_contract': current_contract,
            'available_rooms': available_rooms,
        }
        return render(request, 'dormitory/student_dashboard.html', context)
        
    except Student.DoesNotExist:
        return redirect('complete_profile')
@login_required
def complete_profile(request):
    # Ho√†n thi·ªán h·ªì s∆° sinh vi√™n
    if request.user.user_type != 'student':
        return redirect('home')
    
    if request.method == 'POST':
        student_id = request.POST['student_id']
        university = request.POST['university']
        faculty = request.POST['faculty']
        course = request.POST['course']
        
        # Ki·ªÉm tra xem student_id ƒë√£ t·ªìn t·∫°i ch∆∞a
        if Student.objects.filter(student_id=student_id).exists():
            messages.error(request, "M√£ sinh vi√™n ƒë√£ t·ªìn t·∫°i!")
            return render(request, 'dormitory/complete_profile.html')
        
        Student.objects.create(
            user=request.user,
            student_id=student_id,
            university=university,
            faculty=faculty,
            course=course
        )
        
        messages.success(request, "Ho√†n t·∫•t h·ªì s∆° th√†nh c√¥ng!")
        return redirect('student_dashboard')
    
    return render(request, 'dormitory/complete_profile.html')

def dashboard(request):
    # Ch·ªâ cho ph√©p manager/staff truy c·∫≠p
    # if request.user.user_type not in ['manager', 'staff']:
    #     return render(request, 'errors/access_denied.html')
    
    # Th·ªëng k√™ t·ªïng quan
    total_buildings = Building.objects.count()
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    occupied_rooms = Room.objects.filter(status='occupied').count()
    total_students = Student.objects.count()
    active_contracts = Contract.objects.filter(status='active').count()
    
    # T√≠nh t·ª∑ l·ªá l·∫•p ƒë·∫ßy
    if total_rooms > 0:
        occupancy_percentage = (occupied_rooms / total_rooms) * 100
    else:
        occupancy_percentage = 0
    
    # Ph√≤ng s·∫Øp h·∫øt h·ª£p ƒë·ªìng (trong 30 ng√†y t·ªõi)
    from datetime import date, timedelta
    upcoming_expiry = Contract.objects.filter(
        status='active',
        end_date__lte=date.today() + timedelta(days=30)
    ).count()
    
    # TH√äM PH·∫¶N N√ÄY: TH√îNG B√ÅO THANH TO√ÅN
    from payment.models import Payment
    from django.utils import timezone
    
    today = timezone.now().date()
    
    # H√≥a ƒë∆°n qu√° h·∫°n
    overdue_payments = Payment.objects.filter(
        status='pending', 
        due_date__lt=today
    ).select_related('contract__student', 'contract__room')[:5]  # 5 c√°i g·∫ßn nh·∫•t
    
    # H√≥a ƒë∆°n s·∫Øp ƒë·∫øn h·∫°n (7 ng√†y t·ªõi)
    upcoming_payments = Payment.objects.filter(
        status='pending',
        due_date__range=[today, today + timedelta(days=7)]
    ).select_related('contract__student', 'contract__room')[:5]
    
    # Th·ªëng k√™ thanh to√°n
    total_pending_payments = Payment.objects.filter(status='pending').count()
    total_overdue_payments = Payment.objects.filter(status='pending', due_date__lt=today).count()
    
    context = {
        'stats': {
            'total_buildings': total_buildings,
            'total_rooms': total_rooms,
            'available_rooms': available_rooms,
            'occupied_rooms': occupied_rooms,
            'occupancy_percentage': occupancy_percentage,
            'total_students': total_students,
            'active_contracts': active_contracts,
            'upcoming_expiry': upcoming_expiry,
            # TH√äM TH·ªêNG K√ä THANH TO√ÅN
            'total_pending_payments': total_pending_payments,
            'total_overdue_payments': total_overdue_payments,
        },
        # TH√äM TH√îNG B√ÅO
        'overdue_payments': overdue_payments,
        'upcoming_payments': upcoming_payments,
        'today': today,
    }
    return render(request, 'dormitory/dashboard.html', context)

# dormitory/views.py - TH√äM CU·ªêI FILE
from django.shortcuts import render, get_object_or_404, redirect
from .forms import RoomForm
from django.db.models import Q
# dormitory/views.py - S·ª¨A room_list
from django.core.paginator import Paginator

def room_list(request):
    """Danh s√°ch ph√≤ng v·ªõi t√¨m ki·∫øm v√† ph√¢n trang"""
    rooms = Room.objects.select_related('building', 'room_type').all()
    
    # T√¨m ki·∫øm
    search_query = request.GET.get('search', '')
    if search_query:
        rooms = rooms.filter(
            Q(room_number__icontains=search_query) |
            Q(building__name__icontains=search_query) |
            Q(room_type__name__icontains=search_query)
        )
    
    # Ph√¢n trang - 10 items per page
    paginator = Paginator(rooms, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/room_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })


def room_create(request):
    """Th√™m ph√≤ng m·ªõi"""
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm()
    return render(request, 'dormitory/room_form.html', {'form': form})

def room_update(request, pk):
    """S·ª≠a ph√≤ng"""
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm(instance=room)
    return render(request, 'dormitory/room_form.html', {'form': form})

def room_delete(request, pk):
    """X√≥a ph√≤ng"""
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        room.delete()
        return redirect('room_list')
    return render(request, 'dormitory/room_confirm_delete.html', {'room': room})

# dormitory/views.py - TH√äM CU·ªêI FILE
from .forms import RoomForm, BuildingForm

def building_list(request):
    """Danh s√°ch t√≤a nh√† v·ªõi t√¨m ki·∫øm v√† ph√¢n trang"""
    buildings = Building.objects.all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        buildings = buildings.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    
    # Ph√¢n trang - 10 items per page
    paginator = Paginator(buildings, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/building_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def building_create(request):
    """Th√™m t√≤a nh√† m·ªõi"""
    if request.method == 'POST':
        form = BuildingForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('building_list')
    else:
        form = BuildingForm()
    return render(request, 'dormitory/building_form.html', {'form': form})

def building_update(request, pk):
    """S·ª≠a t√≤a nh√†"""
    building = get_object_or_404(Building, pk=pk)
    if request.method == 'POST':
        form = BuildingForm(request.POST, instance=building)
        if form.is_valid():
            form.save()
            return redirect('building_list')
    else:
        form = BuildingForm(instance=building)
    return render(request, 'dormitory/building_form.html', {'form': form})

def building_delete(request, pk):
    """X√≥a t√≤a nh√†"""
    building = get_object_or_404(Building, pk=pk)
    if request.method == 'POST':
        building.delete()
        return redirect('building_list')
    return render(request, 'dormitory/building_confirm_delete.html', {'building': building})

# dormitory/views.py - TH√äM CU·ªêI FILE
from .forms import RoomForm, BuildingForm, StudentForm

def student_list(request):
    """Danh s√°ch sinh vi√™n v·ªõi t√¨m ki·∫øm v√† ph√¢n trang"""
    students = Student.objects.select_related('user').all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(university__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
    
    # Ph√¢n trang - 10 items per page
    paginator = Paginator(students, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/student_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def student_create(request):
    """Th√™m sinh vi√™n m·ªõi"""
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'dormitory/student_form.html', {'form': form})

def student_update(request, pk):
    """S·ª≠a sinh vi√™n"""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'dormitory/student_form.html', {'form': form})

def student_delete(request, pk):
    """X√≥a sinh vi√™n"""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'dormitory/student_confirm_delete.html', {'student': student})


# dormitory/views.py - TH√äM CU·ªêI FILE
from .forms import RoomForm, BuildingForm, StudentForm, ContractForm

def contract_list(request):
    """Danh s√°ch h·ª£p ƒë·ªìng v·ªõi t√¨m ki·∫øm v√† ph√¢n trang"""
    contracts = Contract.objects.select_related('student', 'room').all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        contracts = contracts.filter(
            Q(contract_number__icontains=search_query) |
            Q(student__student_id__icontains=search_query) |
            Q(student__full_name__icontains=search_query) |
            Q(room__room_number__icontains=search_query) |
            Q(room__building__name__icontains=search_query)
        )
    
    # Ph√¢n trang - 10 items per page
    paginator = Paginator(contracts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/contract_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def contract_create(request):
    """Th√™m h·ª£p ƒë·ªìng m·ªõi"""
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('contract_list')
    else:
        form = ContractForm()
    return render(request, 'dormitory/contract_form.html', {'form': form})

def contract_update(request, pk):
    """S·ª≠a h·ª£p ƒë·ªìng"""
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            form.save()
            return redirect('contract_list')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'dormitory/contract_form.html', {'form': form})

def contract_delete(request, pk):
    """X√≥a h·ª£p ƒë·ªìng"""
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('contract_list')
    return render(request, 'dormitory/contract_confirm_delete.html', {'contract': contract})

# dormitory/views.py
def reports(request):
    """Trang b√°o c√°o th·ªëng k√™"""
    # Th·ªëng k√™ ph√≤ng
    total_rooms = Room.objects.count()
    occupied_rooms = Room.objects.filter(status='occupied').count()
    
    # T√≠nh t·ª∑ l·ªá l·∫•p ƒë·∫ßy
    if total_rooms > 0:
        occupancy_percentage = (occupied_rooms / total_rooms) * 100
    else:
        occupancy_percentage = 0
    
    room_stats = {
        'total': total_rooms,
        'available': Room.objects.filter(status='available').count(),
        'occupied': occupied_rooms,
        'maintenance': Room.objects.filter(status='maintenance').count(),
        'occupancy_percentage': occupancy_percentage,  # Th√™m t·ª∑ l·ªá ph·∫ßn trƒÉm
    }
    
    # Th·ªëng k√™ h·ª£p ƒë·ªìng
    from datetime import date, timedelta
    contract_stats = {
        'active': Contract.objects.filter(status='active').count(),
        'expired': Contract.objects.filter(status='expired').count(),
        'upcoming_expiry': Contract.objects.filter(
            status='active',
            end_date__lte=date.today() + timedelta(days=30)
        ).count(),
    }
    
    # Th·ªëng k√™ theo t√≤a nh√†
    building_stats = []
    for building in Building.objects.all():
        building_total = building.room_set.count()
        building_occupied = building.room_set.filter(status='occupied').count()
        
        # T√≠nh t·ª∑ l·ªá l·∫•p ƒë·∫ßy cho t·ª´ng t√≤a nh√†
        if building_total > 0:
            building_occupancy_rate = (building_occupied / building_total) * 100
        else:
            building_occupancy_rate = 0
        
        building_stats.append({
            'name': building.name,
            'total_rooms': building_total,
            'occupied_rooms': building_occupied,
            'occupancy_rate': building_occupancy_rate,  # T·ª∑ l·ªá ph·∫ßn trƒÉm
            'available_rooms': building_total - building_occupied,  # Ph√≤ng tr·ªëng
        })
    
    context = {
        'room_stats': room_stats,
        'contract_stats': contract_stats,
        'building_stats': building_stats,
    }
    return render(request, 'dormitory/reports.html', context)


import io
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook
from django.utils import timezone
def export_rooms_pdf(request):
    """Xu·∫•t danh s√°ch ph√≤ng PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_phong.pdf"'
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Ti√™u ƒë·ªÅ
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "DANH S√ÅCH PH√íNG - K√ù T√öC X√Å")
    p.setFont("Helvetica", 10)
    p.drawString(100, 730, f"Ng√†y xu·∫•t: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Header table
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, 700, "M√£ ph√≤ng")
    p.drawString(120, 700, "T√≤a nh√†")
    p.drawString(200, 700, "Lo·∫°i ph√≤ng")
    p.drawString(300, 700, "T·∫ßng")
    p.drawString(350, 700, "Tr·∫°ng th√°i")
    
    # D·ªØ li·ªáu
    rooms = Room.objects.select_related('building', 'room_type').all()
    y = 680
    p.setFont("Helvetica", 9)
    
    for room in rooms:
        if y < 100:  # T·∫°o trang m·ªõi n·∫øu h·∫øt ch·ªó
            p.showPage()
            y = 750
            # V·∫Ω header l·∫°i cho trang m·ªõi
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, 750, "M√£ ph√≤ng")
            p.drawString(120, 750, "T√≤a nh√†")
            p.drawString(200, 750, "Lo·∫°i ph√≤ng")
            p.drawString(300, 750, "T·∫ßng")
            p.drawString(350, 750, "Tr·∫°ng th√°i")
            p.setFont("Helvetica", 9)
            y = 730
        
        p.drawString(50, y, room.room_number)
        p.drawString(120, y, room.building.name)
        p.drawString(200, y, room.room_type.name)
        p.drawString(300, y, str(room.floor))
        p.drawString(350, y, room.get_status_display())
        y -= 20
    
    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def export_rooms_excel(request):
    """Xu·∫•t danh s√°ch ph√≤ng Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_phong.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh s√°ch ph√≤ng"
    
    # Header
    headers = ['M√£ ph√≤ng', 'T√≤a nh√†', 'Lo·∫°i ph√≤ng', 'S·ª©c ch·ª©a', 'Gi√° thu√™', 'T·∫ßng', 'Tr·∫°ng th√°i']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # D·ªØ li·ªáu
    rooms = Room.objects.select_related('building', 'room_type').all()
    for row, room in enumerate(rooms, 2):
        ws.cell(row=row, column=1, value=room.room_number)
        ws.cell(row=row, column=2, value=room.building.name)
        ws.cell(row=row, column=3, value=room.room_type.name)
        ws.cell(row=row, column=4, value=room.room_type.capacity)
        ws.cell(row=row, column=5, value=float(room.room_type.price_per_month))
        ws.cell(row=row, column=6, value=room.floor)
        ws.cell(row=row, column=7, value=room.get_status_display())
    
    wb.save(response)
    return response

def export_students_excel(request):
    """Xu·∫•t danh s√°ch sinh vi√™n Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_sinh_vien.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh s√°ch sinh vi√™n"
    
    # Header
    headers = ['M√£ SV', 'H·ªç t√™n', 'Ng√†y sinh', 'Email', 'Tr∆∞·ªùng', 'Khoa', 'Kh√≥a h·ªçc']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # D·ªØ li·ªáu
    students = Student.objects.select_related('user').all()
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.full_name or student.user.get_full_name())
        ws.cell(row=row, column=3, value=student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '')
        ws.cell(row=row, column=4, value=student.user.email)
        ws.cell(row=row, column=5, value=student.university)
        ws.cell(row=row, column=6, value=student.faculty)
        ws.cell(row=row, column=7, value=student.course)
    
    wb.save(response)
    return response

# dormitory/views.py
@login_required
def room_booking(request, room_id):
    """ƒêƒÉng k√Ω ph√≤ng cho sinh vi√™n"""
    if request.user.user_type != 'student':
        messages.error(request, "Ch·ªâ sinh vi√™n m·ªõi c√≥ th·ªÉ ƒëƒÉng k√Ω ph√≤ng!")
        return redirect('home')
    
    room = get_object_or_404(Room, pk=room_id)
    student = request.user.student
    
    # KI·ªÇM TRA PH√íNG C√íN CH·ªñ TR·ªêNG KH√îNG
    if not room.is_available():
        messages.error(request, f"Ph√≤ng {room.room_number} ƒë√£ ƒë·∫ßy!")
        return redirect('student_dashboard')
    
    # Ki·ªÉm tra sinh vi√™n ƒë√£ c√≥ h·ª£p ƒë·ªìng active ch∆∞a
    existing_contract = Contract.objects.filter(student=student, status='active').first()
    if existing_contract:
        messages.warning(request, f"B·∫°n ƒë√£ c√≥ h·ª£p ƒë·ªìng ph√≤ng {existing_contract.room.room_number}!")
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        # T·∫°o h·ª£p ƒë·ªìng m·ªõi
        from datetime import date, timedelta
        
        contract = Contract.objects.create(
            contract_number=f"CT{date.today().strftime('%Y%m%d')}{student.student_id}",
            student=student,
            room=room,
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),  # 1 nƒÉm
            deposit=room.room_type.price_per_month,  # C·ªçc 1 th√°ng
            status='active'
        )
        
        # C·∫¨P NH·∫¨T S·ªê L∆Ø·ª¢NG NG∆Ø·ªúI TRONG PH√íNG
        room.current_occupancy += 1
        
        # KI·ªÇM TRA N·∫æU PH√íNG ƒê√É ƒê·∫¶Y TH√å C·∫¨P NH·∫¨T STATUS
        if room.current_occupancy >= room.room_type.capacity:
            room.status = 'occupied'
        
        room.save()
        
        messages.success(request, f"‚úÖ ƒê√£ ƒëƒÉng k√Ω th√†nh c√¥ng ph√≤ng {room.room_number}!")
        return redirect('student_dashboard')
    
    return render(request, 'dormitory/room_booking.html', {
        'room': room,
        'student': student
    })

from django.contrib.auth import login
from .forms import StudentRegistrationForm

# Trong dormitory/views.py - S·ª¨A H√ÄM student_register
from django.contrib.auth import login
from .forms import StudentRegistrationForm
from .models import Student

# dormitory/views.py - S·ª¨A H√ÄM student_register
def student_register(request):
    """ƒêƒÉng k√Ω t√†i kho·∫£n sinh vi√™n"""
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            try:
                # T·∫°o user (ƒë√£ ƒë∆∞·ª£c x·ª≠ l√Ω trong form.save())
                user = form.save()
                
                # T·∫°o student profile
                student = Student.objects.create(
                    user=user,
                    student_id=form.cleaned_data['student_id'],
                    university=form.cleaned_data['university'],
                    faculty=form.cleaned_data['faculty'],
                    course=form.cleaned_data['course'],
                    full_name=f"{form.cleaned_data['first_name']} {form.cleaned_data['last_name']}",
                    date_of_birth=form.cleaned_data.get('date_of_birth')
                )
                
                messages.success(request, 'üéâ ƒêƒÉng k√Ω t√†i kho·∫£n th√†nh c√¥ng! Vui l√≤ng ƒëƒÉng nh·∫≠p.')
                return redirect('login')
                
            except Exception as e:
                messages.error(request, f'‚ùå C√≥ l·ªói x·∫£y ra: {str(e)}')
                # X√≥a user ƒë√£ t·∫°o n·∫øu c√≥ l·ªói
                if 'user' in locals():
                    user.delete()
        else:
            # Hi·ªÉn th·ªã l·ªói form chi ti·∫øt
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"L·ªói {field}: {error}")
    else:
        form = StudentRegistrationForm()
    
    return render(request, 'dormitory/student_register.html', {'form': form})